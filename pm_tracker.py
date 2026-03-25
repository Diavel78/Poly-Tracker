#!/usr/bin/env python3
"""Polymarket US Position Tracker — Phase 1

Connects to the Polymarket US (regulated DCM) API via the official SDK,
pulls open positions, current market prices, and P&L, then outputs a
formatted terminal summary and saves an XLSX workbook.
"""

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
KEY_ID = os.getenv("POLYMARKET_KEY_ID", "")
SECRET_KEY = os.getenv("POLYMARKET_SECRET_KEY", "")
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# ---------------------------------------------------------------------------
# SDK client
# ---------------------------------------------------------------------------

def get_client(authenticated: bool = True):
    """Return a PolymarketUS client instance."""
    from polymarket_us import PolymarketUS

    if authenticated:
        if not KEY_ID or not SECRET_KEY:
            print("ERROR: POLYMARKET_KEY_ID and POLYMARKET_SECRET_KEY must be set in .env")
            sys.exit(1)
        return PolymarketUS(key_id=KEY_ID, secret_key=SECRET_KEY)
    return PolymarketUS()


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def fetch_positions(client):
    """Fetch all open positions from portfolio endpoint."""
    try:
        response = client.portfolio.positions()
        # The SDK may return a paginated object or list; normalise to list
        if hasattr(response, "data"):
            return list(response.data)
        if hasattr(response, "results"):
            return list(response.results)
        if isinstance(response, list):
            return response
        # If it's an iterable wrapper, try converting
        return list(response)
    except Exception as e:
        print(f"ERROR fetching positions: {e}")
        return []


def fetch_market(client, market_id: str):
    """Fetch a single market by ID."""
    try:
        return client.markets.retrieve(market_id)
    except Exception:
        return None


def fetch_market_price(client, market_slug: str):
    """Fetch best-bid/offer for a market to get the current mid price."""
    try:
        bbo = client.markets.bbo(market_slug)
        best_bid = _safe_float(getattr(bbo, "best_bid_price", None) or getattr(bbo, "bid", None))
        best_ask = _safe_float(getattr(bbo, "best_ask_price", None) or getattr(bbo, "ask", None))
        if best_bid is not None and best_ask is not None:
            return (best_bid + best_ask) / 2
        return best_bid or best_ask
    except Exception:
        return None


def fetch_activities(client):
    """Fetch trade / activity history."""
    try:
        response = client.portfolio.activities()
        if hasattr(response, "data"):
            return list(response.data)
        if hasattr(response, "results"):
            return list(response.results)
        if isinstance(response, list):
            return response
        return list(response)
    except Exception as e:
        print(f"ERROR fetching activities: {e}")
        return []


def fetch_orders(client):
    """Fetch open orders."""
    try:
        response = client.orders.list()
        if hasattr(response, "data"):
            return list(response.data)
        if hasattr(response, "results"):
            return list(response.results)
        if isinstance(response, list):
            return response
        return list(response)
    except Exception as e:
        print(f"ERROR fetching orders: {e}")
        return []


def fetch_balances(client):
    """Fetch account balances."""
    try:
        return client.account.balances()
    except Exception as e:
        print(f"ERROR fetching balances: {e}")
        return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(val):
    """Convert a value to float, returning None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _getattr_chain(obj, *attrs, default=None):
    """Try multiple attribute names, return first non-None value."""
    for attr in attrs:
        val = getattr(obj, attr, None)
        if val is not None:
            return val
    return default


# ---------------------------------------------------------------------------
# Position enrichment
# ---------------------------------------------------------------------------

def enrich_positions(client, positions):
    """Add current price and P&L to each position."""
    enriched = []
    for pos in positions:
        # Extract core fields — attribute names may vary; try common patterns
        market_id = _getattr_chain(pos, "market_id", "marketId", "market")
        market_slug = _getattr_chain(pos, "market_slug", "marketSlug", "slug")
        market_name = _getattr_chain(pos, "market_name", "marketName", "title", "question")
        side = _getattr_chain(pos, "side", "outcome", default="YES")
        quantity = _safe_float(_getattr_chain(pos, "quantity", "size", "qty", "amount")) or 0
        entry_price = _safe_float(_getattr_chain(pos, "entry_price", "entryPrice", "avg_price", "avgPrice"))

        # If we don't have a name yet, fetch the market
        if not market_name and market_id:
            market = fetch_market(client, str(market_id))
            if market:
                market_name = _getattr_chain(market, "title", "question", "name", default=str(market_id))
                if not market_slug:
                    market_slug = _getattr_chain(market, "slug")

        # Get current price
        current_price = None
        if market_slug:
            current_price = fetch_market_price(client, str(market_slug))
        if current_price is None and market_id:
            # Fallback: try market_id as slug
            current_price = fetch_market_price(client, str(market_id))

        # If still no current price, try extracting from position object itself
        if current_price is None:
            current_price = _safe_float(_getattr_chain(pos, "current_price", "currentPrice", "price"))

        # Calculate P&L
        pnl = None
        pnl_pct = None
        current_value = None
        if current_price is not None and quantity:
            current_value = quantity * current_price
            if entry_price is not None:
                pnl = quantity * (current_price - entry_price)
                if entry_price > 0:
                    pnl_pct = ((current_price - entry_price) / entry_price) * 100

        enriched.append({
            "market_name": market_name or str(market_id or "Unknown"),
            "market_id": str(market_id or ""),
            "market_slug": str(market_slug or ""),
            "side": str(side).upper(),
            "quantity": quantity,
            "entry_price": entry_price,
            "current_price": current_price,
            "current_value": current_value,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
        })

    return enriched


# ---------------------------------------------------------------------------
# Terminal output
# ---------------------------------------------------------------------------

def print_positions(enriched, now: datetime):
    """Print a formatted position summary to the terminal."""
    date_str = now.strftime("%Y-%m-%d")
    print()
    print(f"POLYMARKET US POSITIONS — {date_str}")
    print("━" * 90)

    if not enriched:
        print("  No open positions found.")
        print("━" * 90)
        return

    header = f"{'Market':<35} {'Side':<6} {'Qty':>6} {'Entry':>8} {'Current':>8} {'P&L':>10}"
    print(header)
    print("─" * 90)

    total_value = 0.0
    total_pnl = 0.0

    for p in enriched:
        name = p["market_name"][:33]
        side = p["side"][:5]
        qty = f"{p['quantity']:.0f}"
        entry = f"${p['entry_price']:.2f}" if p["entry_price"] is not None else "N/A"
        curr = f"${p['current_price']:.2f}" if p["current_price"] is not None else "N/A"
        if p["pnl"] is not None:
            sign = "+" if p["pnl"] >= 0 else ""
            pnl_str = f"{sign}${p['pnl']:.2f}"
        else:
            pnl_str = "N/A"

        print(f"{name:<35} {side:<6} {qty:>6} {entry:>8} {curr:>8} {pnl_str:>10}")

        if p["current_value"] is not None:
            total_value += p["current_value"]
        if p["pnl"] is not None:
            total_pnl += p["pnl"]

    print("━" * 90)
    sign = "+" if total_pnl >= 0 else ""
    print(f"Total Value: ${total_value:.2f}    Total P&L: {sign}${total_pnl:.2f}")
    print()


# ---------------------------------------------------------------------------
# XLSX output
# ---------------------------------------------------------------------------

def write_xlsx(enriched, activities, now: datetime):
    """Write positions, trade history, and summary to an XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = OUTPUT_DIR / f"polymarket_positions_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")
    bold = Font(bold=True)
    usd_fmt = '"$"#,##0.00'
    pct_fmt = '0.00"%"'

    # ---- Positions tab ----
    ws_pos = wb.active
    ws_pos.title = "Positions"
    pos_headers = ["Market", "Side", "Quantity", "Entry Price", "Current Price",
                   "Unrealized P&L", "% Return"]
    for col, h in enumerate(pos_headers, 1):
        cell = ws_pos.cell(row=1, column=col, value=h)
        cell.font = bold

    for row_idx, p in enumerate(enriched, 2):
        ws_pos.cell(row=row_idx, column=1, value=p["market_name"])
        ws_pos.cell(row=row_idx, column=2, value=p["side"])
        ws_pos.cell(row=row_idx, column=3, value=p["quantity"])

        c_entry = ws_pos.cell(row=row_idx, column=4, value=p["entry_price"])
        c_entry.number_format = usd_fmt

        c_curr = ws_pos.cell(row=row_idx, column=5, value=p["current_price"])
        c_curr.number_format = usd_fmt

        c_pnl = ws_pos.cell(row=row_idx, column=6, value=p["pnl"])
        c_pnl.number_format = usd_fmt
        if p["pnl"] is not None:
            if p["pnl"] >= 0:
                c_pnl.fill = green_fill
                c_pnl.font = green_font
            else:
                c_pnl.fill = red_fill
                c_pnl.font = red_font

        c_pct = ws_pos.cell(row=row_idx, column=7, value=p["pnl_pct"])
        c_pct.number_format = pct_fmt
        if p["pnl_pct"] is not None:
            if p["pnl_pct"] >= 0:
                c_pct.fill = green_fill
                c_pct.font = green_font
            else:
                c_pct.fill = red_fill
                c_pct.font = red_font

    # Auto-width
    for col_cells in ws_pos.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws_pos.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

    # ---- Trade History tab ----
    ws_trades = wb.create_sheet("Trade History")
    trade_headers = ["Timestamp", "Market", "Side", "Price", "Quantity", "Type"]
    for col, h in enumerate(trade_headers, 1):
        cell = ws_trades.cell(row=1, column=col, value=h)
        cell.font = bold

    for row_idx, act in enumerate(activities, 2):
        ts = _getattr_chain(act, "timestamp", "created_at", "createdAt", "time")
        market = _getattr_chain(act, "market_name", "marketName", "title", "market")
        side = _getattr_chain(act, "side", "outcome")
        price = _safe_float(_getattr_chain(act, "price", "fill_price", "fillPrice"))
        qty = _safe_float(_getattr_chain(act, "quantity", "size", "qty", "amount"))
        act_type = _getattr_chain(act, "type", "action", "activity_type", default="trade")

        ws_trades.cell(row=row_idx, column=1, value=str(ts or ""))
        ws_trades.cell(row=row_idx, column=2, value=str(market or ""))
        ws_trades.cell(row=row_idx, column=3, value=str(side or ""))
        c_price = ws_trades.cell(row=row_idx, column=4, value=price)
        c_price.number_format = usd_fmt
        ws_trades.cell(row=row_idx, column=5, value=qty)
        ws_trades.cell(row=row_idx, column=6, value=str(act_type or ""))

    for col_cells in ws_trades.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws_trades.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

    # ---- Summary tab ----
    ws_sum = wb.create_sheet("Summary")
    total_invested = 0.0
    total_current = 0.0
    total_pnl = 0.0
    resolved_wins = 0
    resolved_total = 0

    for p in enriched:
        if p["entry_price"] is not None and p["quantity"]:
            total_invested += p["quantity"] * p["entry_price"]
        if p["current_value"] is not None:
            total_current += p["current_value"]
        if p["pnl"] is not None:
            total_pnl += p["pnl"]
            # Resolved positions have current_price of 0 or 1
            if p["current_price"] in (0.0, 1.0):
                resolved_total += 1
                if p["pnl"] > 0:
                    resolved_wins += 1

    win_rate = (resolved_wins / resolved_total * 100) if resolved_total > 0 else None

    summary_rows = [
        ("Report Date", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Positions", len(enriched)),
        ("Total Invested", total_invested),
        ("Total Current Value", total_current),
        ("Total Unrealized P&L", total_pnl),
        ("Resolved Positions", resolved_total),
        ("Win Rate (resolved)", f"{win_rate:.1f}%" if win_rate is not None else "N/A"),
    ]
    for row_idx, (label, val) in enumerate(summary_rows, 1):
        c_label = ws_sum.cell(row=row_idx, column=1, value=label)
        c_label.font = bold
        c_val = ws_sum.cell(row=row_idx, column=2, value=val)
        if isinstance(val, float):
            c_val.number_format = usd_fmt
            if label == "Total Unrealized P&L":
                if val >= 0:
                    c_val.fill = green_fill
                    c_val.font = green_font
                else:
                    c_val.fill = red_fill
                    c_val.font = red_font

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 20

    wb.save(filename)
    print(f"XLSX saved to: {filename}")
    return filename


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    now = datetime.now(timezone.utc)
    print("Connecting to Polymarket US API...")

    client = get_client(authenticated=True)

    # Fetch data
    print("Fetching positions...")
    positions = fetch_positions(client)
    print(f"  Found {len(positions)} position(s)")

    print("Fetching activity history...")
    activities = fetch_activities(client)
    print(f"  Found {len(activities)} activity record(s)")

    print("Fetching balances...")
    balances = fetch_balances(client)
    if balances is not None:
        print(f"  Balances: {balances}")

    # Enrich positions with current prices and P&L
    print("Enriching positions with current prices...")
    enriched = enrich_positions(client, positions)

    # Terminal output
    print_positions(enriched, now)

    # XLSX output
    print("Writing XLSX report...")
    write_xlsx(enriched, activities, now)

    print("Done.")


if __name__ == "__main__":
    main()
